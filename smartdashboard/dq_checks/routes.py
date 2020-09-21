from flask import render_template, request, url_for, jsonify, Response, Blueprint
from smartdashboard import db, manifest_hive_monitoring, manifest_oracle_monitoring, cdr_threshold
from smartdashboard.utils import init_list, format_date, number_formatter, insert_cdr

from sqlalchemy import or_, and_
from sqlalchemy.sql import func

from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import io, csv
import pprint
dq_blueprint = Blueprint('dq_blueprint', __name__)


@dq_blueprint.route('/dq_overview')
def dq_overview():
    return render_template('dqchecks_overview.html')

# DQ-OVERVIEW HIVE
@dq_blueprint.route('/dqchecks_overview_hive_js', methods=['GET', 'POST'])
def dqchecks_overview_hive_js():
    if request.method == "POST":
        start_date = request.form["start_date"]
        end_date = request.form["end_date"]
        period_select = request.form["period"]
    elif request.method == "GET":
        date_today = date.today()
        start_date = date_today - relativedelta(months=4)
        end_date = date_today
        period_select = "day"

    if period_select == "day":
        period = manifest_hive_monitoring.file_date
    elif period_select == "month":
        period = func.month(manifest_hive_monitoring.file_date)
    elif period_select == "year":
        period = func.year(manifest_hive_monitoring.file_date)
    
    dates = db.session.query(period).filter(and_(manifest_hive_monitoring.file_date >= start_date,manifest_hive_monitoring.file_date <= end_date)).group_by(period).all()
    variances = db.session.query(period,manifest_hive_monitoring.cdr_type,func.sum(manifest_hive_monitoring.variance)).filter(and_(manifest_hive_monitoring.file_date >= start_date,manifest_hive_monitoring.file_date <= end_date)).group_by(period,manifest_hive_monitoring.cdr_type).all()
    len_date = len(dates)

    date_list = init_list(len_date)
    variance_com = init_list(len_date)
    variance_vou = init_list(len_date)
    variance_first = init_list(len_date)
    variance_mon = init_list(len_date)
    variance_cm = init_list(len_date)
    variance_adj = init_list(len_date)
    variance_data = init_list(len_date)
    variance_voice = init_list(len_date)
    variance_sms = init_list(len_date)
    variance_clr = init_list(len_date)

    for i, d in enumerate(dates):
        date_list[i] = format_date(d[0],period_select)
        for v in variances:
            if v.cdr_type == "com" and v[0] == d[0]:
                variance_com[i] = str(v[2])
            elif v.cdr_type == "vou" and v[0] == d[0]:
                variance_vou[i] = str(v[2])
            elif v.cdr_type == "first" and v[0] == d[0]:
                variance_first[i] = str(v[2])
            elif v.cdr_type == "mon" and v[0] == d[0]:
                variance_mon[i] = str(v[2])
            elif v.cdr_type == "cm" and v[0] == d[0]:
                variance_cm[i] = str(v[2])
            elif v.cdr_type == "adj" and v[0] == d[0]:
                variance_adj[i] = str(v[2])
            elif v.cdr_type == "data" and v[0] == d[0]:
                variance_data[i] = str(v[2])
            elif v.cdr_type == "voice" and v[0] == d[0]:
                variance_voice[i] = str(v[2])
            elif v.cdr_type == "sms" and v[0] == d[0]:
                variance_sms[i] = str(v[2])
            elif v.cdr_type == "clr" and v[0] == d[0]:
                variance_clr[i] = str(v[2])

    ### EXPERIMENTAL CODE IN CASE OF SLOWDOWN
    # for i, d in enumerate(dates):
    #     date_list[i] = format_date(d[0],period_select)
    # for v in variances:
    #     if v.cdr_type == "com":
    #         variance_com[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "vou":
    #         variance_vou[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "first":
    #         variance_first[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "mon":
    #         variance_mon[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "cm":
    #         variance_cm[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "adj":
    #         variance_adj[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "data":
    #         variance_data[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "voice":
    #         variance_voice[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "sms":
    #         variance_sms[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    #     elif v.cdr_type == "clr":
    #         variance_clr[date_list.index(format_date(v.file_date,period_select))] = str(v[2])
    
    
    result_set = {
        "date_list": date_list,
        "variance_com": variance_com,
        "variance_vou": variance_vou,
        "variance_first": variance_first,
        "variance_mon": variance_mon,
        "variance_cm": variance_cm,
        "variance_adj": variance_adj,
        "variance_data": variance_data,
        "variance_voice": variance_voice,
        "variance_sms": variance_sms,
        "variance_clr": variance_clr,
        }

    return jsonify(result_set)

@dq_blueprint.route('/dqchecks_hive_excel', methods=['POST'])
def dqchecks_hive_excel():
    if request.method == "POST":
        start_date = request.form["start_date"]
        end_date = request.form["end_date"]
        period_select = request.form["period"]

    if period_select == "day":
        period = manifest_hive_monitoring.file_date
    elif period_select == "month":
        period = func.month(manifest_hive_monitoring.file_date)
    elif period_select == "year":
        period = func.year(manifest_hive_monitoring.file_date)
    
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    dates = db.session.query(period).filter(and_(manifest_hive_monitoring.file_date >= start_date,manifest_hive_monitoring.file_date <= end_date)).group_by(period).all()
    lookup = db.session.query(period,manifest_hive_monitoring.cdr_type,func.sum(manifest_hive_monitoring.ocs_manifest),func.sum(manifest_hive_monitoring.t1_hive),func.sum(manifest_hive_monitoring.variance)).filter(and_(manifest_hive_monitoring.file_date >= start_date,manifest_hive_monitoring.file_date <= end_date)).group_by(period,manifest_hive_monitoring.cdr_type).all()
    len_date = len(dates)
    dates = [ d[0] for d in dates ]

    cdr_dict = {}

    for l in lookup:
        if l.cdr_type not in cdr_dict.keys():
            cdr_dict[l.cdr_type] = {
                "manifest": init_list(len_date),
                "t1": init_list(len_date),
                "variance": init_list(len_date)
            }
            insert_cdr(cdr_dict[l.cdr_type],dates.index(l[0]),l[2],l[3],l[4])
        else:
            insert_cdr(cdr_dict[l.cdr_type],dates.index(l[0]),l[2],l[3],l[4])
    
    print(cdr_dict)
    #output in bytes
    output = io.BytesIO()
    #create WorkBook object
    workbook = Workbook()
    workbook_name = "Manifest Hive Bashing {}".format(datetime.now().strftime("%Y-%m-%d %H-%M-%S"))
    #add a sheet
    ws = workbook.create_sheet('Manifest vs T1 Bashing Validation',0)

    greenFill = PatternFill(start_color='AEEA00',
                   end_color='AEEA00',
                   fill_type='solid')

    x_pos = 1
    y_pos = 1
    temp_y = 1
    x_lim = 19
    row = 0
    
    for c in cdr_dict.keys():
        ### CREATE THE HEADER FOR THE CDR
        # Merge for Date Column
        ws.cell(row=y_pos, column=x_pos, value="Date").alignment = Alignment(horizontal='center')
        # ws.cell(row=y_pos, column=x_pos, value="Date").fill = greenFill
        ws.merge_cells(start_row=y_pos, start_column=x_pos, end_row=y_pos+1, end_column=x_pos)
        # Merge for CDR Row
        ws.cell(row=y_pos, column=x_pos+1, value=c).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=y_pos, start_column=x_pos+1, end_row=y_pos, end_column=x_pos+3)
        ws.cell(row=y_pos+1, column=x_pos+1, value="MANIFEST").alignment = Alignment(horizontal='center')
        ws.cell(row=y_pos+1, column=x_pos+2, value="T1").alignment = Alignment(horizontal='center')
        ws.cell(row=y_pos+1, column=x_pos+3, value="VARIANCE").alignment = Alignment(horizontal='center')
        for i, d in enumerate(dates):
            ws.cell(row=y_pos+2, column=x_pos, value=d.strftime("%m/%d/%y"))
            ws.cell(row=y_pos+2, column=x_pos+1, value=cdr_dict[c]["manifest"][i])
            ws.cell(row=y_pos+2, column=x_pos+2, value=cdr_dict[c]["t1"][i])
            ws.cell(row=y_pos+2, column=x_pos+3, value=cdr_dict[c]["variance"][i])
            y_pos+=1

        ### SET X AND Y POSITIONS
        if x_pos+3 < 19:
            y_pos = temp_y
            x_pos += 5
        else:
            row += 1
            y_pos = temp_y + (len(dates)+3)
            temp_y = y_pos
            x_pos = 1
    
    workbook.save(output)
    output.seek(0)
    pprint.pprint(cdr_dict)
    filename = workbook_name

    return Response(output, mimetype="application/openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment;filename={}.xlsx".format(filename)})

@dq_blueprint.route('/dqchecks_hive_table', methods=['POST'])
def dqchecks_hive_table():
    start_date = request.form["start_date"]
    end_date = request.form["end_date"]

    if start_date == "" and end_date == "":
        date_today = date.today()
        start_date = date_today - relativedelta(days=8)
        end_date = date_today
    
    lookup = db.session.query(manifest_hive_monitoring.file_date,
                                manifest_hive_monitoring.cdr_type,
                                func.sum(manifest_hive_monitoring.ocs_manifest),
                                func.sum(manifest_hive_monitoring.t1_hive),
                                func.sum(manifest_hive_monitoring.variance))\
                                .filter(and_(manifest_hive_monitoring.file_date >= start_date,manifest_hive_monitoring.file_date <= end_date))\
                                .group_by(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type)

    data = []
    for l in lookup:
        data.append({
            "file date": l.file_date.strftime("%Y-%m-%d"),
            "cdr": l.cdr_type,
            "manifest count": str(number_formatter(l[2])),
            "t1 count": str(number_formatter(l[3])),
            "variance": str(number_formatter(l[4]))
        })
    result_set = {
        "data": data
    }

    return jsonify(result_set)

# DQ-OVERVIEW ORACLE
@dq_blueprint.route('/dqchecks_overview_oracle_js', methods=['GET', 'POST'])
def dqchecks_overview_oracle_js():
    if request.method == "POST":
        start_date = request.form["start_date"]
        end_date = request.form["end_date"]
        period_select = request.form["period"]
    elif request.method == "GET":
        date_today = date.today()
        start_date = date_today - relativedelta(months=4)
        end_date = date_today
        period_select = "day"

    if period_select == "day":
        period = manifest_oracle_monitoring.file_date
    elif period_select == "month":
        period = func.month(manifest_oracle_monitoring.file_date)
    elif period_select == "year":
        period = func.year(manifest_oracle_monitoring.file_date)
    
    dates = db.session.query(period).filter(and_(manifest_oracle_monitoring.file_date >= start_date,manifest_oracle_monitoring.file_date <= end_date)).group_by(period).all()
    variances = db.session.query(period,manifest_oracle_monitoring.cdr_type,func.sum(manifest_oracle_monitoring.variance)).filter(and_(manifest_oracle_monitoring.file_date >= start_date,manifest_oracle_monitoring.file_date <= end_date)).group_by(period,manifest_oracle_monitoring.cdr_type).all()
    len_date = len(dates)

    date_list = init_list(len_date)
    variance_com = init_list(len_date)
    variance_vou = init_list(len_date)
    variance_first = init_list(len_date)
    variance_mon = init_list(len_date)
    variance_cm = init_list(len_date)
    variance_adj = init_list(len_date)
    variance_data = init_list(len_date)
    variance_voice = init_list(len_date)
    variance_sms = init_list(len_date)
    variance_clr = init_list(len_date)


    for i, d in enumerate(dates):
        date_list[i] = format_date(d[0],period_select)
        for v in variances:
            if v.cdr_type == "com" and v[0] == d[0]:
                variance_com[i] = str(v[2])
            elif v.cdr_type == "vou" and v[0] == d[0]:
                variance_vou[i] = str(v[2])
            elif v.cdr_type == "first" and v[0] == d[0]:
                variance_first[i] = str(v[2])
            elif v.cdr_type == "mon" and v[0] == d[0]:
                variance_mon[i] = str(v[2])
            elif v.cdr_type == "cm" and v[0] == d[0]:
                variance_cm[i] = str(v[2])
            elif v.cdr_type == "adj" and v[0] == d[0]:
                variance_adj[i] = str(v[2])
            elif v.cdr_type == "data" and v[0] == d[0]:
                variance_data[i] = str(v[2])
            elif v.cdr_type == "voice" and v[0] == d[0]:
                variance_voice[i] = str(v[2])
            elif v.cdr_type == "sms" and v[0] == d[0]:
                variance_sms[i] = str(v[2])
            elif v.cdr_type == "clr" and v[0] == d[0]:
                variance_clr[i] = str(v[2])
    
    
    result_set = {
        "date_list": date_list,
        "variance_com": variance_com,
        "variance_vou": variance_vou,
        "variance_first": variance_first,
        "variance_mon": variance_mon,
        "variance_cm": variance_cm,
        "variance_adj": variance_adj,
        "variance_data": variance_data,
        "variance_voice": variance_voice,
        "variance_sms": variance_sms,
        "variance_clr": variance_clr,
        }

    return jsonify(result_set)

@dq_blueprint.route('/dqchecks_oracle_table', methods=['POST'])
def dqchecks_oracle_table():
    start_date = request.form["start_date"]
    end_date = request.form["end_date"]

    if start_date == "" and end_date == "":
        date_today = date.today()
        start_date = date_today - relativedelta(days=8)
        end_date = date_today
   
    lookup = db.session.query(manifest_oracle_monitoring.file_date,
                                manifest_oracle_monitoring.cdr_type,
                                func.sum(manifest_oracle_monitoring.ocs_manifest),
                                func.sum(manifest_oracle_monitoring.t1_oracle),
                                func.sum(manifest_oracle_monitoring.variance))\
                                .filter(and_(manifest_oracle_monitoring.file_date >= start_date,manifest_oracle_monitoring.file_date <= end_date))\
                                .group_by(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type)

    data = []
    for l in lookup:
        data.append({
            "file date": l.file_date.strftime("%Y-%m-%d"),
            "cdr": l.cdr_type,
            "manifest count": str(number_formatter(l[2])),
            "t1 count": str(number_formatter(l[3])),
            "variance": str(number_formatter(l[4]))
        })
    result_set = {
        "data": data
    }

    return jsonify(result_set)

@dq_blueprint.route('/dqchecks_exce_oracle_excel', methods=['POST'])
def dqchecks_exce_oracle_excel():
    if request.method == "POST":
        start_date = request.form["start_date"]
        end_date = request.form["end_date"]
        period_select = request.form["period"]

    if period_select == "day":
        period = manifest_oracle_monitoring.file_date
    elif period_select == "month":
        period = func.month(manifest_oracle_monitoring.file_date)
    elif period_select == "year":
        period = func.year(manifest_oracle_monitoring.file_date)
    
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    dates = db.session.query(period).filter(and_(manifest_oracle_monitoring.file_date >= start_date,manifest_oracle_monitoring.file_date <= end_date)).group_by(period).all()
    lookup = db.session.query(period,manifest_oracle_monitoring.cdr_type,func.sum(manifest_oracle_monitoring.ocs_manifest),func.sum(manifest_oracle_monitoring.t1_oracle),func.sum(manifest_oracle_monitoring.variance)).filter(and_(manifest_oracle_monitoring.file_date >= start_date,manifest_oracle_monitoring.file_date <= end_date)).group_by(period,manifest_oracle_monitoring.cdr_type).all()
    len_date = len(dates)
    dates = [ d[0] for d in dates ]

    cdr_dict = {}

    for l in lookup:
        if l.cdr_type not in cdr_dict.keys():
            cdr_dict[l.cdr_type] = {
                "manifest": init_list(len_date),
                "t1": init_list(len_date),
                "variance": init_list(len_date)
            }
            insert_cdr(cdr_dict[l.cdr_type],dates.index(l[0]),l[2],l[3],l[4])
        else:
            insert_cdr(cdr_dict[l.cdr_type],dates.index(l[0]),l[2],l[3],l[4])

# MANIFEST VS HIVE
@dq_blueprint.route('/dq_manvshive')
def dq_manvshive():
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    lookup = db.session.query(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type,func.sum(manifest_hive_monitoring.variance)).filter(manifest_hive_monitoring.file_date == date.today()).group_by(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type)
    variance_dict = dict.fromkeys(cdr_types,None)
    
    for l in lookup:
        if l.cdr_type not in variance_dict.keys():
            variance_dict[l.cdr_type] = l[2]
        else:
            variance_dict[l.cdr_type] = l[2]

    return render_template('dqchecks_manvshive.html', variance_com = number_formatter(variance_dict["com"]),
                                                variance_vou = number_formatter(variance_dict["vou"]),
                                                variance_first = number_formatter(variance_dict["first"]),
                                                variance_mon = number_formatter(variance_dict["mon"]),
                                                variance_cm = number_formatter(variance_dict["cm"]),
                                                variance_adj = number_formatter(variance_dict["adj"]),
                                                variance_data = number_formatter(variance_dict["data"]),
                                                variance_voice = number_formatter(variance_dict["voice"]),
                                                variance_sms = number_formatter(variance_dict["sms"]),
                                                variance_clr = number_formatter(variance_dict["clr"])
                                                )

@dq_blueprint.route('/dqchecks_manvshive_js', methods=['GET','POST'])
def dqchecks_manvshive_js():
    if request.method == "POST":
        query_date = request.form["cdr_date"]
    else:
        query_date = date.today()

    results = db.session.query(manifest_hive_monitoring).filter(manifest_hive_monitoring.file_date == query_date)
    com_manifest = init_list()
    com_t1 = init_list()
    com_variance = init_list()
    vou_manifest = init_list()
    vou_t1 = init_list()
    vou_variance = init_list()
    first_manifest = init_list()
    first_t1 = init_list()
    first_variance = init_list()
    mon_manifest = init_list()
    mon_t1 = init_list()
    mon_variance = init_list()
    cm_manifest = init_list()
    cm_t1 = init_list()
    cm_variance = init_list()
    adj_manifest = init_list()
    adj_t1 = init_list()
    adj_variance = init_list()
    data_manifest = init_list()
    data_t1 = init_list()
    data_variance = init_list()
    voice_manifest = init_list()
    voice_t1 = init_list()
    voice_variance = init_list()
    sms_manifest = init_list()
    sms_t1 = init_list()
    sms_variance = init_list()
    clr_manifest = init_list()
    clr_t1 = init_list()
    clr_variance = init_list()


    for r in results:
        if r.cdr_type == "com":
            com_manifest[r.processing_hour] = (str(r.ocs_manifest))
            com_t1[r.processing_hour] = (str(r.t1_hive))
            com_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "vou":
            vou_manifest[r.processing_hour] = (str(r.ocs_manifest))
            vou_t1[r.processing_hour] = (str(r.t1_hive))
            vou_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "first":
            first_manifest[r.processing_hour] = (str(r.ocs_manifest))
            first_t1[r.processing_hour] = (str(r.t1_hive))
            first_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "mon":
            mon_manifest[r.processing_hour] = (str(r.ocs_manifest))
            mon_t1[r.processing_hour] = (str(r.t1_hive))
            mon_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "cm":
            cm_manifest[r.processing_hour] = (str(r.ocs_manifest))
            cm_t1[r.processing_hour] = (str(r.t1_hive))
            cm_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "adj":
            adj_manifest[r.processing_hour] = (str(r.ocs_manifest))
            adj_t1[r.processing_hour] = (str(r.t1_hive))
            adj_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "data":
            data_manifest[r.processing_hour] = (str(r.ocs_manifest))
            data_t1[r.processing_hour] = (str(r.t1_hive))
            data_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "voice":
            voice_manifest[r.processing_hour] = (str(r.ocs_manifest))
            voice_t1[r.processing_hour] = (str(r.t1_hive))
            voice_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "sms":
            sms_manifest[r.processing_hour] = (str(r.ocs_manifest))
            sms_t1[r.processing_hour] = (str(r.t1_hive))
            sms_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "clr":
            clr_manifest[r.processing_hour] = (str(r.ocs_manifest))
            clr_t1[r.processing_hour] = (str(r.t1_hive))
            clr_variance[r.processing_hour] = (str(r.variance))

    result_set = {
        "com_manifest": com_manifest,
        "com_t1": com_t1,
        "com_variance": com_variance,
        "vou_manifest": vou_manifest,
        "vou_t1": vou_t1,
        "vou_variance": vou_variance,
        "first_manifest": first_manifest,
        "first_t1": first_t1,
        "first_variance": first_variance,
        "mon_manifest": mon_manifest,
        "mon_t1": mon_t1,
        "mon_variance": mon_variance,
        "cm_manifest": cm_manifest,
        "cm_t1": cm_t1,
        "cm_variance": cm_variance,
        "adj_manifest": adj_manifest,
        "adj_t1": adj_t1,
        "adj_variance": adj_variance,
        "data_manifest": data_manifest,
        "data_t1": data_t1,
        "data_variance": data_variance,
        "voice_manifest": voice_manifest,
        "voice_t1": voice_t1,
        "voice_variance": voice_variance,
        "sms_manifest": sms_manifest,
        "sms_t1": sms_t1,
        "sms_variance": sms_variance,
        "clr_manifest": clr_manifest,
        "clr_t1": clr_t1,
        "clr_variance": clr_variance
    }

    return jsonify(result_set)

@dq_blueprint.route('/dqchecks_manvshive_search', methods=['POST'])
def dqchecks_manvshive_search(): 
    if request.method == "POST":
        query_date = request.form["date"]
        cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
        lookup = db.session.query(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type,func.sum(manifest_hive_monitoring.variance)).filter(manifest_hive_monitoring.file_date == query_date).group_by(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type)
        variance_dict = dict.fromkeys(cdr_types,None)
        for l in lookup:
            if l.cdr_type not in variance_dict.keys():
                variance_dict[l.cdr_type] = l[2]
            else:
                variance_dict[l.cdr_type] = l[2]
      
    result_set = {
        "variance_com": number_formatter(variance_dict["com"]),
        "variance_vou": number_formatter(variance_dict["vou"]),
        "variance_first": number_formatter(variance_dict["first"]),
        "variance_mon": number_formatter(variance_dict["mon"]),
        "variance_cm": number_formatter(variance_dict["cm"]),
        "variance_adj": number_formatter(variance_dict["adj"]),
        "variance_data": number_formatter(variance_dict["data"]),
        "variance_voice": number_formatter(variance_dict["voice"]),
        "variance_sms": number_formatter(variance_dict["sms"]),
        "variance_clr": number_formatter(variance_dict["clr"])
    }

    return jsonify(result_set)

# MANIFEST VS ORACLE
@dq_blueprint.route('/dq_manvsoracle', methods=['GET', 'POST'])
def dq_manvsoracle():
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    lookup = db.session.query(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type,func.sum(manifest_oracle_monitoring.variance)).filter(manifest_oracle_monitoring.file_date == date.today()).group_by(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type)
    variance_dict = dict.fromkeys(cdr_types,None)
    
    for l in lookup:
        if l.cdr_type not in variance_dict.keys():
            variance_dict[l.cdr_type] = l[2]
        else:
            variance_dict[l.cdr_type] = l[2]
 
    return render_template('dqchecks_manvsoracle.html', variance_com = number_formatter(variance_dict["com"]),
                                                variance_vou = number_formatter(variance_dict["vou"]),
                                                variance_first = number_formatter(variance_dict["first"]),
                                                variance_mon = number_formatter(variance_dict["mon"]),
                                                variance_cm = number_formatter(variance_dict["cm"]),
                                                variance_adj = number_formatter(variance_dict["adj"]),
                                                variance_data = number_formatter(variance_dict["data"]),
                                                variance_voice = number_formatter(variance_dict["voice"]),
                                                variance_sms = number_formatter(variance_dict["sms"]),
                                                variance_clr = number_formatter(variance_dict["clr"])
                                                )

@dq_blueprint.route('/dqchecks_manvsoracle_js', methods=['GET','POST'])
def dqchecks_manvsoracle_js():
    if request.method == "POST":
        query_date = request.form["cdr_date"]
    else:
        query_date = date.today()
    
    results = db.session.query(manifest_oracle_monitoring).filter(manifest_oracle_monitoring.file_date == query_date)
    com_manifest = init_list()
    com_t1 = init_list()
    com_variance = init_list()
    vou_manifest = init_list()
    vou_t1 = init_list()
    vou_variance = init_list()
    first_manifest = init_list()
    first_t1 = init_list()
    first_variance = init_list()
    mon_manifest = init_list()
    mon_t1 = init_list()
    mon_variance = init_list()
    cm_manifest = init_list()
    cm_t1 = init_list()
    cm_variance = init_list()
    adj_manifest = init_list()
    adj_t1 = init_list()
    adj_variance = init_list()
    data_manifest = init_list()
    data_t1 = init_list()
    data_variance = init_list()
    voice_manifest = init_list()
    voice_t1 = init_list()
    voice_variance = init_list()
    sms_manifest = init_list()
    sms_t1 = init_list()
    sms_variance = init_list()
    clr_manifest = init_list()
    clr_t1 = init_list()
    clr_variance = init_list()

    for r in results:
        if r.cdr_type == "com":
            com_manifest[r.processing_hour] = (str(r.ocs_manifest))
            com_t1[r.processing_hour] = (str(r.t1_oracle))
            com_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "vou":
            vou_manifest[r.processing_hour] = (str(r.ocs_manifest))
            vou_t1[r.processing_hour] = (str(r.t1_oracle))
            vou_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "first":
            first_manifest[r.processing_hour] = (str(r.ocs_manifest))
            first_t1[r.processing_hour] = (str(r.t1_oracle))
            first_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "mon":
            mon_manifest[r.processing_hour] = (str(r.ocs_manifest))
            mon_t1[r.processing_hour] = (str(r.t1_oracle))
            mon_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "cm":
            cm_manifest[r.processing_hour] = (str(r.ocs_manifest))
            cm_t1[r.processing_hour] = (str(r.t1_oracle))
            cm_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "adj":
            adj_manifest[r.processing_hour] = (str(r.ocs_manifest))
            adj_t1[r.processing_hour] = (str(r.t1_oracle))
            adj_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "data":
            data_manifest[r.processing_hour] = (str(r.ocs_manifest))
            data_t1[r.processing_hour] = (str(r.t1_oracle))
            data_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "voice":
            voice_manifest[r.processing_hour] = (str(r.ocs_manifest))
            voice_t1[r.processing_hour] = (str(r.t1_oracle))
            voice_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "sms":
            sms_manifest[r.processing_hour] = (str(r.ocs_manifest))
            sms_t1[r.processing_hour] = (str(r.t1_oracle))
            sms_variance[r.processing_hour] = (str(r.variance))
        elif r.cdr_type == "clr":
            clr_manifest[r.processing_hour] = (str(r.ocs_manifest))
            clr_t1[r.processing_hour] = (str(r.t1_oracle))
            clr_variance[r.processing_hour] = (str(r.variance))

    result_set = {
        "com_manifest": com_manifest,
        "com_t1": com_t1,
        "com_variance": com_variance,
        "vou_manifest": vou_manifest,
        "vou_t1": vou_t1,
        "vou_variance": vou_variance,
        "first_manifest": first_manifest,
        "first_t1": first_t1,
        "first_variance": first_variance,
        "mon_manifest": mon_manifest,
        "mon_t1": mon_t1,
        "mon_variance": mon_variance,
        "cm_manifest": cm_manifest,
        "cm_t1": cm_t1,
        "cm_variance": cm_variance,
        "adj_manifest": adj_manifest,
        "adj_t1": adj_t1,
        "adj_variance": adj_variance,
        "data_manifest": data_manifest,
        "data_t1": data_t1,
        "data_variance": data_variance,
        "voice_manifest": voice_manifest,
        "voice_t1": voice_t1,
        "voice_variance": voice_variance,
        "sms_manifest": sms_manifest,
        "sms_t1": sms_t1,
        "sms_variance": sms_variance,
        "clr_manifest": clr_manifest,
        "clr_t1": clr_t1,
        "clr_variance": clr_variance
    }

    return jsonify(result_set)

@dq_blueprint.route('/dqchecks_manvsoracle_search', methods=['GET','POST'])
def dqchecks_manvsoracle_search():
    if request.method == "POST":
        query_date = request.form["date"]
        cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
        lookup = db.session.query(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type,func.sum(manifest_oracle_monitoring.variance)).filter(manifest_oracle_monitoring.file_date == query_date).group_by(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type)
        variance_dict = dict.fromkeys(cdr_types,None)
        
        for l in lookup:
            if l.cdr_type not in variance_dict.keys():
                variance_dict[l.cdr_type] = l[2]
            else:
                variance_dict[l.cdr_type] = l[2]
    
    result_set = {
        "variance_com": number_formatter(variance_dict["com"]),
        "variance_vou": number_formatter(variance_dict["vou"]),
        "variance_first": number_formatter(variance_dict["first"]),
        "variance_mon": number_formatter(variance_dict["mon"]),
        "variance_cm": number_formatter(variance_dict["cm"]),
        "variance_adj": number_formatter(variance_dict["adj"]),
        "variance_data": number_formatter(variance_dict["data"]),
        "variance_voice": number_formatter(variance_dict["voice"]),
        "variance_sms": number_formatter(variance_dict["sms"]),
        "variance_clr": number_formatter(variance_dict["clr"])
    }

    return jsonify(result_set)

#THRESHOLD
@dq_blueprint.route('/dqchecks_update_threshold', methods=['GET','POST'])
def dqchecks_update_threshold():
    if request.method == "POST":
        cdr_types = request.form["cdrs"]
        custom_threshold = request.form["threshold"]

        # CREATING adding new cdr type and threshold
        # lookup = cdr_threshold(cdr_type = cdr_types, threshold = custom_threshold )
        # db.session.add(lookup)
        # db.session.commit()

        #UPDATING threshold according to cdr
        lookup = db.session.query(cdr_threshold).filter(cdr_threshold.cdr_type == cdr_types).first()
        lookup.cdr_type = cdr_types
        lookup.threshold = custom_threshold
        db.session.commit()


    threshold_cdrs = {}
    threshold = db.session.query(cdr_threshold).all()

    for t in threshold:
        threshold_cdrs[t.cdr_type] = str(number_formatter(t.threshold))

    return jsonify(threshold_cdrs)


# FOR LZERO
@dq_blueprint.route('/dq_overview_lzero')
def dq_overview_lzero():
    return render_template('dqchecks_overview_lzero.html')

@dq_blueprint.route('/dq_manvshive_lzero')
def dq_manvshive_lzero():
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    lookup = db.session.query(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type,func.sum(manifest_hive_monitoring.variance)).filter(manifest_hive_monitoring.file_date == date.today()).group_by(manifest_hive_monitoring.file_date,manifest_hive_monitoring.cdr_type)
    variance_dict = dict.fromkeys(cdr_types,None)
    
    for l in lookup:
        if l.cdr_type not in variance_dict.keys():
            variance_dict[l.cdr_type] = l[2]
        else:
            variance_dict[l.cdr_type] = l[2]

    return render_template('dqchecks_manvshive_lzero.html', variance_com = number_formatter(variance_dict["com"]),
                                                variance_vou = number_formatter(variance_dict["vou"]),
                                                variance_first = number_formatter(variance_dict["first"]),
                                                variance_mon = number_formatter(variance_dict["mon"]),
                                                variance_cm = number_formatter(variance_dict["cm"]),
                                                variance_adj = number_formatter(variance_dict["adj"]),
                                                variance_data = number_formatter(variance_dict["data"]),
                                                variance_voice = number_formatter(variance_dict["voice"]),
                                                variance_sms = number_formatter(variance_dict["sms"]),
                                                variance_clr = number_formatter(variance_dict["clr"])
                                                )

@dq_blueprint.route('/dq_manvsoracle_lzero', methods=['GET', 'POST'])
def dq_manvsoracle_lzero():
    cdr_types = ["com", "vou", "cm", "adj", "first", "mon", "data", "voice", "sms", "clr"]
    lookup = db.session.query(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type,func.sum(manifest_oracle_monitoring.variance)).filter(manifest_oracle_monitoring.file_date == date.today()).group_by(manifest_oracle_monitoring.file_date,manifest_oracle_monitoring.cdr_type)
    variance_dict = dict.fromkeys(cdr_types,None)
    
    for l in lookup:
        if l.cdr_type not in variance_dict.keys():
            variance_dict[l.cdr_type] = l[2]
        else:
            variance_dict[l.cdr_type] = l[2]
 
    return render_template('dqchecks_manvsoracle_lzero.html', variance_com = number_formatter(variance_dict["com"]),
                                                variance_vou = number_formatter(variance_dict["vou"]),
                                                variance_first = number_formatter(variance_dict["first"]),
                                                variance_mon = number_formatter(variance_dict["mon"]),
                                                variance_cm = number_formatter(variance_dict["cm"]),
                                                variance_adj = number_formatter(variance_dict["adj"]),
                                                variance_data = number_formatter(variance_dict["data"]),
                                                variance_voice = number_formatter(variance_dict["voice"]),
                                                variance_sms = number_formatter(variance_dict["sms"]),
                                                variance_clr = number_formatter(variance_dict["clr"])
                                                )
